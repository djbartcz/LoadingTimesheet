import os
import logging
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import load_workbook, Workbook

class ExcelClient:
    def __init__(self, file_path: str):
        """
        Initialize Excel client with local file path
        
        Args:
            file_path: Path to Excel file (e.g., "C:/Users/YourName/OneDrive/TimesheetData.xlsx")
        """
        self.file_path = Path(file_path)
        self.workbook = None
        
    def _load_workbook(self, force_reload=True):
        """Load workbook, create if doesn't exist - always reloads fresh data"""
        try:
            # Close existing workbook if open to ensure fresh reload
            if self.workbook is not None and force_reload:
                try:
                    self.workbook.close()
                except:
                    pass
                self.workbook = None
            
            if self.file_path.exists():
                # Always reload from disk to get latest changes from Excel file
                # This ensures any external edits are immediately visible
                self.workbook = load_workbook(self.file_path)
                logging.debug(f"Reloaded workbook from: {self.file_path}")
            else:
                # Create new workbook if it doesn't exist
                self.workbook = Workbook()
                self.workbook.remove(self.workbook.active)  # Remove default sheet
                self.workbook.save(self.file_path)
                logging.info(f"Created new Excel file: {self.file_path}")
        except Exception as e:
            logging.error(f"Error loading workbook: {e}")
            raise
    
    def _save_workbook(self):
        """Save workbook to file"""
        try:
            self.workbook.save(self.file_path)
            logging.debug(f"Workbook saved: {self.file_path}")
        except Exception as e:
            logging.error(f"Error saving workbook: {e}")
            raise
    
    def get_or_create_worksheet(self, worksheet_name: str, headers: List[str]):
        """Get or create worksheet with headers - reloads workbook first"""
        # Force reload workbook to ensure we have latest data
        # This ensures any external changes are immediately visible
        self._load_workbook()
        
        if worksheet_name not in self.workbook.sheetnames:
            worksheet = self.workbook.create_sheet(worksheet_name)
            worksheet.append(headers)
            self._save_workbook()
            logging.info(f"Created worksheet: {worksheet_name}")
        
        return self.workbook[worksheet_name]
    
    def get_worksheet_data(self, worksheet_name: str) -> List[Dict]:
        """Read all data from worksheet - reloads workbook each time for live data"""
        # Force reload workbook to get latest data from Excel file
        # This ensures any external changes to the Excel file are immediately visible
        self._load_workbook()
        
        if worksheet_name not in self.workbook.sheetnames:
            logging.warning(f"Worksheet '{worksheet_name}' not found")
            return []
        
        worksheet = self.workbook[worksheet_name]
        
        # Get headers from first row
        headers = []
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                headers.append(cell.value if cell.value else "")
        
        if not headers:
            return []
        
        # Get data rows
        records = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip() for cell in row if cell is not None):
                record = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    record[header] = value
                records.append(record)
        
        return records
    
    def append_row(self, worksheet_name: str, row_data: List):
        """
        Append a row to worksheet (automatically finds first free row)
        Reloads workbook before appending to ensure we're working with latest data
        
        Args:
            worksheet_name: Name of the worksheet
            row_data: List of values to append
        """
        # Reload workbook to ensure we have latest data before appending
        self._load_workbook()
        
        worksheet = self.get_or_create_worksheet(worksheet_name, [])
        
        # Append row - openpyxl automatically finds the next free row
        worksheet.append(row_data)
        self._save_workbook()
        
        logging.info(f"Appended row to '{worksheet_name}': {row_data}")
        return True
    
    def find_first_free_row(self, worksheet_name: str) -> int:
        """
        Find first empty row in worksheet (helper method)
        Note: append_row() handles this automatically, but this can be useful for custom logic
        """
        if not self.workbook:
            self._load_workbook()
        
        if worksheet_name not in self.workbook.sheetnames:
            return 2  # Start after header row
        
        worksheet = self.workbook[worksheet_name]
        
        # Find first completely empty row
        for row_num in range(2, worksheet.max_row + 2):
            row = worksheet[row_num]
            if all(cell.value is None or str(cell.value).strip() == '' for cell in row):
                return row_num
        
        return worksheet.max_row + 1

# Global instance (will be initialized in server.py)
excel_client = None

def init_excel_client(file_path: str):
    """Initialize global Excel client"""
    global excel_client
    excel_client = ExcelClient(file_path)
    logging.info(f"Excel client initialized with file: {file_path}")

