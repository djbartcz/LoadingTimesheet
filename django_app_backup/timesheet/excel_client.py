import os
import logging
from pathlib import Path
from typing import List, Dict, Optional
from openpyxl import load_workbook, Workbook
import tempfile
import requests
from urllib.parse import urlparse, quote
import time


class SharePointAuth:
    """Handle SharePoint authentication using Azure AD Client Credentials"""

    def __init__(self, tenant_id: str, client_id: str, client_secret: str,
                 sharepoint_site: str):
        """
        Initialize SharePoint authentication

        Args:
            tenant_id: Azure AD Tenant ID
            client_id: Azure AD App Registration Client ID
            client_secret: Azure AD App Registration Client Secret
            sharepoint_site: SharePoint site URL
        """
        self.tenant_id = tenant_id
        self.client_id = client_id
        self.client_secret = client_secret
        self.sharepoint_site = sharepoint_site
        self.access_token = None
        self.token_expires_at = 0

    def get_access_token(self) -> str:
        """Get or refresh access token"""
        # Check if token is still valid (with 5 minute buffer)
        if (self.access_token and
                time.time() < self.token_expires_at - 300):
            return self.access_token

        # Get new token
        token_url = (
            f"https://login.microsoftonline.com/"
            f"{self.tenant_id}/oauth2/v2.0/token"
        )

        token_data = {
            'client_id': self.client_id,
            'client_secret': self.client_secret,
            'scope': f'{self.sharepoint_site}/.default',
            'grant_type': 'client_credentials'
        }

        try:
            response = requests.post(token_url, data=token_data)
            response.raise_for_status()
            token_response = response.json()

            self.access_token = token_response['access_token']
            expires_in = token_response.get('expires_in', 3600)
            self.token_expires_at = time.time() + expires_in

            logging.info("Successfully obtained SharePoint access token")
            return self.access_token

        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to obtain access token: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logging.error(f"Response: {e.response.text}")
            raise


class ExcelClient:
    def __init__(self, file_path: str,
                 sharepoint_auth: Optional[SharePointAuth] = None):
        """
        Initialize Excel client with local file path or SharePoint URL

        Args:
            file_path: Path to Excel file (local path or SharePoint URL)
            sharepoint_auth: Optional SharePointAuth object for auth
        """
        self.original_path = file_path
        self.is_sharepoint_url = (
            file_path.startswith('http://') or
            file_path.startswith('https://')
        )

        if self.is_sharepoint_url:
            # For SharePoint URLs, we'll download to a temp file
            self.temp_file = None
            self.sharepoint_url = file_path
            self.file_path = None  # Will be set when we download
            self.sharepoint_auth = sharepoint_auth

            # Try to initialize auth from environment if not provided
            if not self.sharepoint_auth:
                self.sharepoint_auth = self._init_auth_from_env()

            if self.sharepoint_auth:
                logging.info(
                    f"Initialized ExcelClient with SharePoint URL "
                    f"(authenticated): {file_path}"
                )
            else:
                logging.warning(
                    f"Initialized ExcelClient with SharePoint URL "
                    f"(no authentication): {file_path}"
                )
        else:
            # Local file path
            self.file_path = Path(file_path)
            self.temp_file = None
            self.sharepoint_url = None
            self.sharepoint_auth = None
            logging.info(
                f"Initialized ExcelClient with local path: {file_path}"
            )

        self.workbook = None

    def _init_auth_from_env(self) -> Optional[SharePointAuth]:
        """Initialize SharePoint auth from environment variables"""
        tenant_id = os.environ.get('SHAREPOINT_TENANT_ID', '').strip()
        client_id = os.environ.get('SHAREPOINT_CLIENT_ID', '').strip()
        client_secret = os.environ.get(
            'SHAREPOINT_CLIENT_SECRET', ''
        ).strip()
        sharepoint_site = os.environ.get('SHAREPOINT_SITE', '').strip()

        if all([tenant_id, client_id, client_secret]):
            # Extract site from URL if not provided
            if not sharepoint_site and self.sharepoint_url:
                parsed = urlparse(self.sharepoint_url)
                sharepoint_site = f"{parsed.scheme}://{parsed.netloc}"

            if sharepoint_site:
                try:
                    return SharePointAuth(
                        tenant_id, client_id, client_secret, sharepoint_site
                    )
                except Exception as e:
                    logging.error(
                        f"Failed to initialize SharePoint auth: {e}"
                    )
                    return None

        return None

    def _convert_sharepoint_url_to_api(self, url: str) -> str:
        """Convert SharePoint web URL to REST API URL"""
        # Parse the URL
        parsed = urlparse(url)

        # Extract site and file path
        # URL format: https://tenant.sharepoint.com/sites/SiteName/...
        path_parts = parsed.path.split('/')
        site_index = None
        for i, part in enumerate(path_parts):
            if part == 'sites' and i + 1 < len(path_parts):
                site_index = i
                break

        if site_index:
            site_name = path_parts[site_index + 1]
            # Get everything after /sites/SiteName/
            file_path = '/'.join(path_parts[site_index + 2:])
            # Remove query parameters from file path
            file_path = file_path.split('?')[0]
            # URL encode the path
            file_path = '/'.join(
                quote(part, safe='') for part in file_path.split('/')
            )

            api_url = (
                f"{parsed.scheme}://{parsed.netloc}/sites/{site_name}/"
                f"_api/web/GetFileByServerRelativeUrl('{file_path}')/$value"
            )
            return api_url
        else:
            # Fallback: try direct download URL
            return url.split('?')[0] + '?download=1'

    def _download_from_sharepoint(self) -> Path:
        """Download file from SharePoint to temporary location"""
        try:
            # Create a temporary file
            temp_fd, temp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(temp_fd)  # Close the file descriptor

            # Prepare headers
            headers = {}
            if self.sharepoint_auth:
                # Use authenticated request
                access_token = self.sharepoint_auth.get_access_token()
                headers['Authorization'] = f'Bearer {access_token}'
                # Convert to API URL for better compatibility
                api_url = self._convert_sharepoint_url_to_api(
                    self.sharepoint_url
                )
                download_url = api_url
            else:
                # Try direct download (may not work without auth)
                download_url = self.sharepoint_url
                logging.warning(
                    "Downloading without authentication - may fail"
                )

            # Download the file
            response = requests.get(
                download_url, headers=headers, stream=True
            )
            response.raise_for_status()

            with open(temp_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)

            self.temp_file = temp_path
            self.file_path = Path(temp_path)
            logging.info(
                f"Downloaded SharePoint file to temp location: {temp_path}"
            )
            return self.file_path

        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to download from SharePoint: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logging.error(
                    f"Response status: {e.response.status_code}"
                )
                logging.error(f"Response text: {e.response.text[:500]}")
            if not self.sharepoint_auth:
                logging.error(
                    "Note: SharePoint URL access requires authentication."
                )
                logging.error("Options:")
                logging.error(
                    "1. Set SHAREPOINT_TENANT_ID, SHAREPOINT_CLIENT_ID, "
                    "and SHAREPOINT_CLIENT_SECRET in .env file"
                )
                logging.error(
                    "2. Sync SharePoint folder locally via OneDrive "
                    "and use local path"
                )
            raise

    def _upload_to_sharepoint(self, local_path: Path):
        """Upload file back to SharePoint"""
        if not self.sharepoint_auth:
            raise Exception(
                "Cannot upload to SharePoint without authentication"
            )

        try:
            # Convert to API URL for upload
            api_url = self._convert_sharepoint_url_to_api(
                self.sharepoint_url
            )
            upload_url = api_url

            # Get access token
            access_token = self.sharepoint_auth.get_access_token()
            headers = {
                'Authorization': f'Bearer {access_token}',
                'Content-Type': (
                    'application/vnd.openxmlformats-officedocument.'
                    'spreadsheetml.sheet'
                )
            }

            # Read file and upload
            with open(local_path, 'rb') as f:
                file_content = f.read()

            response = requests.put(
                upload_url, data=file_content, headers=headers
            )
            response.raise_for_status()

            logging.info(
                f"Uploaded file back to SharePoint: {self.sharepoint_url}"
            )
        except requests.exceptions.RequestException as e:
            logging.error(f"Failed to upload to SharePoint: {e}")
            if hasattr(e, 'response') and e.response is not None:
                logging.error(
                    f"Response status: {e.response.status_code}"
                )
                logging.error(f"Response text: {e.response.text[:500]}")
            logging.warning(
                "File was modified locally but not uploaded to SharePoint"
            )
            raise

    def _load_workbook(self, force_reload=True):
        """Load workbook, create if doesn't exist - reloads fresh data"""
        try:
            # Close existing workbook if open to ensure fresh reload
            if self.workbook is not None and force_reload:
                try:
                    self.workbook.close()
                except Exception:
                    pass
                self.workbook = None

            # Handle SharePoint URLs
            if self.is_sharepoint_url:
                if (self.temp_file is None or
                        not Path(self.temp_file).exists()):
                    # Download from SharePoint
                    self._download_from_sharepoint()
                else:
                    # Re-download to get latest changes
                    self._download_from_sharepoint()

            if self.file_path.exists():
                # Always reload from disk to get latest changes
                # This ensures any external edits are immediately visible
                self.workbook = load_workbook(self.file_path)
                logging.debug(f"Reloaded workbook from: {self.file_path}")
            else:
                # Create new workbook if it doesn't exist
                self.workbook = Workbook()
                self.workbook.remove(self.workbook.active)
                self._save_workbook()
                logging.info(f"Created new Excel file: {self.file_path}")
        except Exception as e:
            logging.error(f"Error loading workbook: {e}")
            raise

    def _save_workbook(self):
        """Save workbook to file"""
        try:
            self.workbook.save(self.file_path)
            logging.debug(f"Workbook saved: {self.file_path}")

            # If it's a SharePoint URL, upload the changes back
            if self.is_sharepoint_url and self.temp_file:
                self._upload_to_sharepoint(self.file_path)
        except Exception as e:
            logging.error(f"Error saving workbook: {e}")
            raise

    def __del__(self):
        """Cleanup temporary file when object is destroyed"""
        if self.temp_file and Path(self.temp_file).exists():
            try:
                os.unlink(self.temp_file)
                logging.debug(
                    f"Cleaned up temporary file: {self.temp_file}"
                )
            except Exception as e:
                logging.warning(f"Failed to cleanup temporary file: {e}")

    def get_or_create_worksheet(
            self, worksheet_name: str, headers: List[str]):
        """Get or create worksheet with headers - reloads workbook first"""
        # Force reload workbook to ensure we have latest data
        # This ensures any external changes are immediately visible
        self._load_workbook()

        if worksheet_name not in self.workbook.sheetnames:
            worksheet = self.workbook.create_sheet(worksheet_name)
            worksheet.append(headers)
            self._save_workbook()
            logging.info(f"Created worksheet: {worksheet_name}")

        return self.workbook[worksheet_name]

    def get_worksheet_data(self, worksheet_name: str) -> List[Dict]:
        """Read all data from worksheet - reloads workbook each time"""
        # Force reload workbook to get latest data from Excel file
        # This ensures any external changes are immediately visible
        self._load_workbook()

        if worksheet_name not in self.workbook.sheetnames:
            logging.warning(f"Worksheet '{worksheet_name}' not found")
            return []

        worksheet = self.workbook[worksheet_name]

        # Get headers from first row
        headers = []
        if worksheet.max_row > 0:
            for cell in worksheet[1]:
                headers.append(cell.value if cell.value else "")

        if not headers:
            return []

        # Get data rows
        records = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            # Skip completely empty rows
            if any(cell is not None and str(cell).strip()
                   for cell in row if cell is not None):
                record = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    record[header] = value
                records.append(record)

        return records

    def append_row(self, worksheet_name: str, row_data: List):
        """
        Append a row to worksheet (automatically finds first free row)
        Reloads workbook before appending to ensure we're working with
        latest data

        Args:
            worksheet_name: Name of the worksheet
            row_data: List of values to append
        """
        # Reload workbook to ensure we have latest data before appending
        self._load_workbook()

        worksheet = self.get_or_create_worksheet(worksheet_name, [])

        # Append row - openpyxl automatically finds the next free row
        worksheet.append(row_data)
        self._save_workbook()

        logging.info(f"Appended row to '{worksheet_name}': {row_data}")
        return True

    def replace_worksheet_data(
            self, worksheet_name: str, headers: List[str], rows: List[List]
    ):
        """
        Replace all data in a worksheet (keeps headers, clears all rows,
        then repopulates). This is useful for syncing data from database
        to Excel

        Args:
            worksheet_name: Name of the worksheet
            headers: List of header names
            rows: List of rows, where each row is a list of values
        """
        # Reload workbook to ensure we have latest data
        self._load_workbook()

        # Get or create worksheet with headers
        worksheet = self.get_or_create_worksheet(worksheet_name, headers)

        # Clear all existing data rows (keep header row)
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row)

        # Append all rows
        for row_data in rows:
            worksheet.append(row_data)

        self._save_workbook()

        logging.info(f"Replaced data in '{worksheet_name}': {len(rows)} rows")
        return True
