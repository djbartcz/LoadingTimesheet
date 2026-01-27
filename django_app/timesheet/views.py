from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.utils import timezone
from django.conf import settings
from datetime import timedelta, datetime
import logging
import uuid
import os
from functools import wraps
import pytz
import json
from .models import ActiveTimer, TimeRecord, Employee, Project, Task
from django.apps import apps
from openpyxl import Workbook
from io import BytesIO
from django.db import transaction
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required as django_login_required
from django.contrib.auth.models import User, Group

logger = logging.getLogger(__name__)


def get_excel_timezone():
    """
    Get the timezone to use for Excel exports.
    Returns system local timezone if EXCEL_TIMEZONE is 'system',
    otherwise uses the configured timezone.
    """
    excel_tz = getattr(settings, 'EXCEL_TIMEZONE', settings.TIME_ZONE)
    
    if excel_tz.lower() == 'system':
        # Use system local timezone
        # Get local timezone from system
        from datetime import datetime, timezone as dt_timezone
        local_tz = datetime.now(dt_timezone.utc).astimezone().tzinfo
        # Convert to pytz timezone if needed
        if isinstance(local_tz, pytz.BaseTzInfo):
            return local_tz
        else:
            # Try to get timezone name and convert
            tz_name = str(local_tz)
            # Common Windows timezone mappings
            tz_mapping = {
                'Central European Standard Time': 'Europe/Prague',
                'Central European Time': 'Europe/Prague',
                'Central Europe Standard Time': 'Europe/Prague',
                'W. Europe Standard Time': 'Europe/Berlin',
                'GMT Standard Time': 'Europe/London',
                'Eastern Standard Time': 'America/New_York',
            }
            if tz_name in tz_mapping:
                return pytz.timezone(tz_mapping[tz_name])
            # Default to Django TIME_ZONE if can't determine
            logger.warning(
                f"Could not determine system timezone '{tz_name}', "
                f"using Django TIME_ZONE"
            )
            return pytz.timezone(settings.TIME_ZONE)
    else:
        # Use configured timezone
        try:
            return pytz.timezone(excel_tz)
        except pytz.exceptions.UnknownTimeZoneError:
            logger.warning(
                f"Unknown timezone '{excel_tz}', using Django TIME_ZONE"
            )
            return pytz.timezone(settings.TIME_ZONE)


def convert_to_excel_timezone(dt):
    """
    Convert a timezone-aware datetime to the Excel timezone for display.
    If datetime is naive, assumes it's in Django's TIME_ZONE.
    """
    if timezone.is_naive(dt):
        # Make naive datetime timezone-aware using Django's TIME_ZONE
        dt = timezone.make_aware(dt, pytz.timezone(settings.TIME_ZONE))
    
    # Convert to Excel timezone
    excel_tz = get_excel_timezone()
    return dt.astimezone(excel_tz)


# Admin required decorator
def admin_required(view_func):
    """
    Decorator to require admin access for views.
    Users must be in the 'Admin' group, have is_staff=True, or be superuser.
    """
    @wraps(view_func)
    @django_login_required
    def wrapper(request, *args, **kwargs):
        # Check if user has admin access (Admin group, staff, or superuser)
        has_admin_access = (
            request.user.groups.filter(name='Admin').exists() or
            request.user.is_staff or
            request.user.is_superuser
        )
        
        if has_admin_access:
            return view_func(request, *args, **kwargs)
        
        from django.urls import reverse
        login_url = reverse('login')
        current_path = request.get_full_path()
        return redirect(f"{login_url}?next={current_path}")
    return wrapper

# Login required decorator (for regular users)
# Just use Django's built-in decorator - it will use LOGIN_URL from settings
login_required = django_login_required

# Get Excel client from app config
def get_excel_client():
    timesheet_config = apps.get_app_config('timesheet')
    return getattr(timesheet_config, 'excel_client', None)


def get_today_start():
    now = timezone.now()
    return now.replace(hour=0, minute=0, second=0, microsecond=0)


def get_week_start():
    now = timezone.now()
    start = now - timedelta(days=now.weekday())
    return start.replace(hour=0, minute=0, second=0, microsecond=0)


def format_duration(seconds):
    """Format seconds to human readable duration"""
    if not seconds:
        return '0min'
    hrs = seconds // 3600
    mins = (seconds % 3600) // 60
    if hrs > 0:
        return f"{hrs}h {mins}min"
    return f"{mins}min"


def format_time(seconds):
    """Format seconds to HH:MM:SS"""
    hrs = seconds // 3600
    mins = (seconds % 3600) // 60
    secs = seconds % 60
    return f"{hrs:02d}:{mins:02d}:{secs:02d}"


# Login View
def login_view(request):
    """Login page - handles authentication using Django User model"""
    # If already authenticated, redirect to home
    if request.user.is_authenticated:
        return redirect('employee_selection')
    
    # Handle POST request (login form submission)
    if request.method == 'POST':
        username = request.POST.get('username', '')
        password = request.POST.get('password', '')
        
        # Authenticate using Django's authentication system
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            # Login the user
            login(request, user)
            # Redirect to next page if specified, otherwise to employee selection
            next_url = request.GET.get('next', 'employee_selection')
            # Validate next_url to prevent open redirects
            if next_url and not next_url.startswith('http'):
                return redirect(next_url)
            return redirect('employee_selection')
        else:
            return render(request, 'timesheet/login.html', {
                'error': 'Nesprávné přihlašovací údaje'
            })
    
    # GET request - show login form
    return render(request, 'timesheet/login.html')


# Logout View
def logout_view(request):
    """Logout - uses Django logout"""
    logout(request)
    return redirect('login')


@login_required
def employee_selection(request):
    """Main page - employee selection"""
    employees = []
    active_timers = {}
    error_message = None
    try:
        db_employees = Employee.objects.filter(is_active=True).order_by('name')
        employees = [
            {"id": emp.id, "name": emp.name}
            for emp in db_employees
        ]
        if not employees:
            error_message = "No employees found. Please add employees in the Admin Control Panel."
    except Exception as e:
        logger.error(f"Error fetching employees from database: {e}", exc_info=True)
        error_message = f"Error loading employees: {str(e)}"
    
    # Get active timers
    try:
        timers = ActiveTimer.objects.all()
        for timer in timers:
            active_timers[timer.employee_id] = {
                'id': timer.id,
                'employee_id': timer.employee_id,
                'employee_name': timer.employee_name,
                'project_id': timer.project_id,
                'project_name': timer.project_name,
                'task': timer.task,
                'is_non_productive': timer.is_non_productive,
                'is_break': timer.is_break,
                'start_time': timer.start_time.isoformat(),
            }
    except Exception as e:
        logger.error(f"Error fetching active timers: {e}", exc_info=True)
    
    is_admin = (
        request.user.groups.filter(name='Admin').exists() or
        request.user.is_staff or
        request.user.is_superuser
    )
    
    return render(request, 'timesheet/employee_selection.html', {
        'employees': employees,
        'active_timers': active_timers,
        'error_message': error_message,
        'is_admin': is_admin,
    })


@login_required
def timer_page(request, employee_id):
    """Timer page for specific employee"""
    try:
        employee_obj = Employee.objects.get(id=employee_id, is_active=True)
        employee = {"id": employee_obj.id, "name": employee_obj.name}
    except Employee.DoesNotExist:
        logger.warning(f"Employee {employee_id} not found or inactive")
        return redirect('employee_selection')
    except Exception as e:
        logger.error(f"Error fetching employee: {e}")
        return redirect('employee_selection')
    
    projects = []
    tasks = []
    non_productive_tasks = []
    
    try:
        db_projects = Project.objects.filter(is_active=True).order_by('name')
        projects = [
            {
                "id": proj.id, 
                "name": f"{proj.name} - {proj.project_description}" if proj.project_description else proj.name
            }
            for proj in db_projects
        ]
        
        db_tasks = Task.objects.filter(is_active=True, is_non_productive=False).order_by('name')
        tasks = [{"name": task.name} for task in db_tasks]
        
        db_non_prod_tasks = Task.objects.filter(is_active=True, is_non_productive=True).order_by('name')
        non_productive_tasks = [{"name": task.name} for task in db_non_prod_tasks]
    except Exception as e:
        logger.error(f"Error fetching projects/tasks from database: {e}", exc_info=True)
    
    active_timer = None
    try:
        timer = ActiveTimer.objects.get(employee_id=employee_id)
        active_timer = {
            'id': timer.id,
            'employee_id': timer.employee_id,
            'employee_name': timer.employee_name,
            'project_id': timer.project_id,
            'project_name': timer.project_name,
            'task': timer.task,
            'is_non_productive': timer.is_non_productive,
            'is_break': timer.is_break,
            'start_time': timer.start_time.isoformat(),
        }
    except ActiveTimer.DoesNotExist:
        pass
    
    # Get last task
    last_task = None
    try:
        record = TimeRecord.objects.filter(employee_id=employee_id, end_time__isnull=False).order_by('-end_time').first()
        if record:
            last_task = {
                'id': record.id,
                'employee_id': record.employee_id,
                'employee_name': record.employee_name,
                'project_id': record.project_id,
                'project_name': record.project_name,
                'task': record.task,
                'is_non_productive': record.is_non_productive,
                'start_time': record.start_time.isoformat(),
                'end_time': record.end_time.isoformat() if record.end_time else None,
                'duration_seconds': record.duration_seconds,
            }
    except Exception as e:
        logger.error(f"Error fetching last task: {e}")
    
    # Get summary
    today_start = get_today_start()
    week_start = get_week_start()
    
    today_records = TimeRecord.objects.filter(
        employee_id=employee_id,
        end_time__gte=today_start
    ).order_by('-end_time')[:100]
    
    week_records = TimeRecord.objects.filter(
        employee_id=employee_id,
        end_time__gte=week_start
    ).order_by('-end_time')[:500]
    
    today_prod = sum(r.duration_seconds or 0 for r in today_records if not r.is_non_productive)
    today_nonprod = sum(r.duration_seconds or 0 for r in today_records if r.is_non_productive)
    week_prod = sum(r.duration_seconds or 0 for r in week_records if not r.is_non_productive)
    week_nonprod = sum(r.duration_seconds or 0 for r in week_records if r.is_non_productive)
    
    summary = {
        'today': {
            'total_seconds': today_prod + today_nonprod,
            'productive_seconds': today_prod,
            'non_productive_seconds': today_nonprod,
            'break_seconds': 0,
            'records': [
                {
                    'id': r.id,
                    'task': r.task,
                    'project_name': r.project_name,
                    'is_non_productive': r.is_non_productive,
                    'duration_seconds': r.duration_seconds,
                }
                for r in today_records[:10]
            ]
        },
        'week': {
            'total_seconds': week_prod + week_nonprod,
            'productive_seconds': week_prod,
            'non_productive_seconds': week_nonprod,
            'break_seconds': 0,
        }
    }
    
    # Check if this is an AJAX request for timer sync
    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        timer_data = None
        if active_timer:
            timer_data = {
                'id': active_timer['id'],
                'start_time': active_timer['start_time'],
                'is_non_productive': active_timer['is_non_productive'],
                'is_break': active_timer['is_break'],
                'employee_id': active_timer.get('employee_id', employee_id),
            }
        return JsonResponse({
            'has_timer': active_timer is not None,
            'timer': timer_data,
        })
    
    return render(request, 'timesheet/timer_page.html', {
        'employee': employee,
        'projects': projects,
        'tasks': [{'name': t['name']} for t in tasks],
        'non_productive_tasks': [{'name': t['name']} for t in non_productive_tasks],
        'active_timer': active_timer,
        'last_task': last_task,
        'summary': summary,
    })


@login_required
def start_timer(request, employee_id):
    """Start a new timer"""
    if request.method != 'POST':
        return redirect('timer_page', employee_id=employee_id)
    
    try:
        try:
            employee_obj = Employee.objects.get(id=employee_id, is_active=True)
            employee = {"id": employee_obj.id, "name": employee_obj.name}
        except Employee.DoesNotExist:
            logger.warning(f"Employee {employee_id} not found or inactive")
            return redirect('employee_selection')
        except Exception as e:
            logger.error(f"Error fetching employee: {e}")
            return redirect('employee_selection')
        
        mode = request.POST.get('mode', 'productive')
        project_id = request.POST.get('project_id', '')
        project_name = request.POST.get('project_name', '')
        task = request.POST.get('task', '')
        non_productive_task = request.POST.get('non_productive_task', '')
        
        if ActiveTimer.objects.filter(employee_id=employee_id).exists():
            return redirect('timer_page', employee_id=employee_id)
        
        if mode == 'productive':
            if not project_id or not task:
                return redirect('timer_page', employee_id=employee_id)
            is_non_productive = False
            final_task = task
        else:
            if not non_productive_task:
                return redirect('timer_page', employee_id=employee_id)
            is_non_productive = True
            final_task = non_productive_task
            project_id = None
            project_name = None
        
        start_time = timezone.now()
        record = ActiveTimer(
            id=str(uuid.uuid4()),
            employee_id=employee_id,
            employee_name=employee['name'],
            project_id=project_id,
            project_name=project_name,
            task=final_task,
            is_non_productive=is_non_productive,
            is_break=False,
            start_time=start_time,  # Already timezone-aware from timezone.now()
        )
        record.save()
        
        return redirect('timer_page', employee_id=employee_id)
    except Exception as e:
        logger.error(f"Error starting timer: {e}")
        return redirect('timer_page', employee_id=employee_id)


@login_required
def stop_timer(request, employee_id):
    """Stop a timer and save to database, then Excel - handles form submission"""
    if request.method != 'POST':
        return redirect('timer_page', employee_id=employee_id)
    
    try:
        # Get active timer
        timer = get_object_or_404(ActiveTimer, employee_id=employee_id)
        
        # Use timezone-aware datetime for accurate calculation
        end_time = timezone.now()
        
        # Ensure both times are timezone-aware for accurate calculation
        if timezone.is_naive(timer.start_time):
            start_time = timezone.make_aware(timer.start_time)
        else:
            start_time = timer.start_time
        
        # Calculate duration accurately
        time_delta = end_time - start_time
        duration_seconds = int(time_delta.total_seconds())
        
        # Validate duration (should be positive)
        if duration_seconds < 0:
            logger.warning(f"Negative duration detected for timer {timer.id}, using 0")
            duration_seconds = 0
        
        # Format duration
        hours = duration_seconds // 3600
        minutes = (duration_seconds % 3600) // 60
        seconds = duration_seconds % 60
        duration_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
        
        # Save to database FIRST (this is the source of truth)
        # Database is primary storage - ensures data is never lost
        time_record = TimeRecord(
            id=timer.id,
            employee_id=timer.employee_id,
            employee_name=timer.employee_name,
            project_id=timer.project_id,
            project_name=timer.project_name,
            task=timer.task,
            is_non_productive=timer.is_non_productive,
            start_time=start_time,  # Use timezone-aware start_time
            end_time=end_time,
            duration_seconds=duration_seconds,
        )
        time_record.save()
        logger.info(f"✓ Saved time record to database: {timer.employee_name} - {duration_seconds}s ({duration_formatted})")
        
        # Then save to Excel (for reporting/export)
        # Excel is secondary storage - if this fails, data is still safe in database
        excel_client = get_excel_client()
        if excel_client:
            try:
                # Convert to Excel timezone (local/system time) for display
                start_dt_excel = convert_to_excel_timezone(start_time)
                end_dt_excel = convert_to_excel_timezone(end_time)
                
                # Convert seconds to hours (decimal)
                duration_hours = round(duration_seconds / 3600.0, 2)
                
                if timer.is_non_productive:
                    excel_client.get_or_create_worksheet("Neproduktivní záznamy", [
                        "Datum", "Zaměstnanec ID", "Zaměstnanec", "Úkon",
                        "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"
                    ])
                    row = [
                        start_dt_excel.strftime("%Y-%m-%d"),
                        timer.employee_id,
                        timer.employee_name,
                        timer.task,
                        start_dt_excel.strftime("%H:%M:%S"),
                        end_dt_excel.strftime("%H:%M:%S"),
                        duration_formatted,
                        duration_hours
                    ]
                    excel_client.append_row("Neproduktivní záznamy", row)
                    logger.info(f"✓ Exported to Excel: Neproduktivní záznamy")
                else:
                    excel_client.get_or_create_worksheet("Záznamy", [
                        "Datum", "Zaměstnanec ID", "Zaměstnanec", "Projekt ID", "Projekt",
                        "Úkon", "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"
                    ])
                    row = [
                        start_dt_excel.strftime("%Y-%m-%d"),
                        timer.employee_id,
                        timer.employee_name,
                        timer.project_id or '',
                        timer.project_name or '',
                        timer.task,
                        start_dt_excel.strftime("%H:%M:%S"),
                        end_dt_excel.strftime("%H:%M:%S"),
                        duration_formatted,
                        duration_hours
                    ]
                    excel_client.append_row("Záznamy", row)
                    logger.info(f"✓ Exported to Excel: Záznamy")
            except Exception as excel_error:
                logger.error(f"⚠ Error exporting to Excel (record already saved to DB): {excel_error}")
                # Don't fail if Excel export fails - data is already in database
        
        # Delete active timer
        timer.delete()
        
        return redirect('timer_page', employee_id=employee_id)
    except Exception as e:
        logger.error(f"Error stopping timer: {e}")
        return redirect('timer_page', employee_id=employee_id)


@admin_required
def admin_dashboard(request):
    """Admin dashboard"""
    employees = []
    try:
        db_employees = Employee.objects.filter(is_active=True).order_by('name')
        employees = [
            {"id": emp.id, "name": emp.name}
            for emp in db_employees
            if emp and emp.id and emp.name
        ]
    except Exception as e:
        logger.error(f"Error fetching employees for admin dashboard: {e}", exc_info=True)
        employees = []
    
    # Ensure employees is a list
    if not isinstance(employees, list):
        employees = []
    
    today_start = get_today_start()
    week_start = get_week_start()
    now = timezone.now()
    month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    
    try:
        active_timers = list(ActiveTimer.objects.all())
        today_records = list(TimeRecord.objects.filter(end_time__gte=today_start)[:1000])
        week_records = list(TimeRecord.objects.filter(end_time__gte=week_start)[:5000])
        month_records_for_stats = list(TimeRecord.objects.filter(end_time__gte=month_start)[:10000])
        
        employee_stats = []
        active_map = {}
        for t in active_timers:
            if t and t.employee_id:
                active_map[t.employee_id] = t
        
        for emp in employees:
            if not emp or not isinstance(emp, dict) or 'id' not in emp:
                continue
            emp_id = emp.get('id')
            if not emp_id:
                continue
            emp_today = [r for r in today_records if r and r.employee_id == emp_id]
            emp_week = [r for r in week_records if r and r.employee_id == emp_id]
            emp_month = [r for r in month_records_for_stats if r and r.employee_id == emp_id]
            
            today_secs = sum((r.duration_seconds or 0) for r in emp_today)
            week_secs = sum((r.duration_seconds or 0) for r in emp_week)
            month_secs = sum((r.duration_seconds or 0) for r in emp_month)
            
            active = active_map.get(emp_id)
            
            employee_stats.append({
                'employee_id': emp_id,
                'employee_name': emp.get('name', 'Unknown'),
                'today_seconds': today_secs,
                'today_hours': round(today_secs / 3600, 2),
                'week_seconds': week_secs,
                'week_hours': round(week_secs / 3600, 2),
                'month_seconds': month_secs,
                'month_hours': round(month_secs / 3600, 2),
                'is_working': active is not None,
                'is_on_break': False,
                'is_non_productive': getattr(active, 'is_non_productive', False) if active else False,
                'current_task': getattr(active, 'task', None) if active else None,
                'current_project': getattr(active, 'project_name', None) if active else None,
                'start_time': active.start_time.isoformat() if active and hasattr(active, 'start_time') and active.start_time else None,
            })
    except Exception as e:
        logger.error(f"Error fetching time records for admin dashboard: {e}", exc_info=True)
        employee_stats = []
    
    # Sort employee stats safely
    try:
        employee_stats.sort(key=lambda x: (not x.get('is_working', False), x.get('employee_name', '')))
    except Exception as e:
        logger.error(f"Error sorting employee stats: {e}", exc_info=True)
    
    # Ensure employee_stats is a list
    if not employee_stats:
        employee_stats = []
    
    total_today = sum(s.get('today_seconds', 0) or 0 for s in employee_stats)
    total_week = sum(s.get('week_seconds', 0) or 0 for s in employee_stats)
    working_count = sum(1 for s in employee_stats if s.get('is_working', False))
    
    # Use month_records_for_stats if available, otherwise fetch again
    try:
        if 'month_records_for_stats' in locals():
            month_records = month_records_for_stats
        else:
            month_records = list(TimeRecord.objects.filter(end_time__gte=month_start)[:10000])
    except Exception as e:
        logger.error(f"Error fetching month records: {e}", exc_info=True)
        month_records = []
    
    # Statistics per person (for charts)
    stats_per_person = []
    if employee_stats:
        for emp in employee_stats:
            if emp and isinstance(emp, dict) and 'employee_id' in emp:
                emp_month = [r for r in month_records if r.employee_id == emp.get('employee_id')]
                month_secs = sum(r.duration_seconds or 0 for r in emp_month)
                prod_secs = sum(r.duration_seconds or 0 for r in emp_month if not r.is_non_productive)
                nonprod_secs = sum(r.duration_seconds or 0 for r in emp_month if r.is_non_productive)
                
                stats_per_person.append({
                    'name': emp.get('employee_name', 'Unknown'),
                    'today_hours': round((emp.get('today_seconds', 0) or 0) / 3600, 2),
                    'week_hours': round((emp.get('week_seconds', 0) or 0) / 3600, 2),
                    'month_hours': round(month_secs / 3600, 2),
                    'productive_hours': round(prod_secs / 3600, 2),
                    'non_productive_hours': round(nonprod_secs / 3600, 2),
                })
    
    # Statistics per project (grouped by project description, with BOMs)
    stats_per_project = {}
    stats_per_bom = {}
    project_cache = {}
    bom_to_project_map = {}
    
    try:
        for record in month_records:
            if record and record.project_id and not record.is_non_productive:
                project_id = record.project_id
                
                if project_id not in project_cache:
                    try:
                        project = Project.objects.get(id=project_id)
                        project_desc = project.project_description or None
                        project_cache[project_id] = {
                            'id': project.id,
                            'name': project.name,
                            'description': project_desc
                        }
                        bom_to_project_map[project_id] = project_desc
                    except Project.DoesNotExist:
                        project_cache[project_id] = {
                            'id': project_id,
                            'name': project_id,
                            'description': record.project_name or None
                        }
                        bom_to_project_map[project_id] = record.project_name or None
                
                project_info = project_cache[project_id]
                project_desc = project_info['description'] or 'Unnamed Project'
                
                hours = (record.duration_seconds or 0) / 3600
                
                if project_desc not in stats_per_project:
                    stats_per_project[project_desc] = {
                        'description': project_desc,
                        'boms': {},
                        'total_hours': 0,
                        'total_records': 0,
                        'employees': set()
                    }
                
                if project_id not in stats_per_project[project_desc]['boms']:
                    stats_per_project[project_desc]['boms'][project_id] = {
                        'bom': project_id,
                        'hours': 0,
                        'records': 0,
                        'employees': set()
                    }
                
                stats_per_project[project_desc]['boms'][project_id]['hours'] += hours
                stats_per_project[project_desc]['boms'][project_id]['records'] += 1
                stats_per_project[project_desc]['boms'][project_id]['employees'].add(record.employee_id)
                
                stats_per_project[project_desc]['total_hours'] += hours
                stats_per_project[project_desc]['total_records'] += 1
                stats_per_project[project_desc]['employees'].add(record.employee_id)
        
        stats_per_project_list = []
        for project_desc, data in stats_per_project.items():
            boms_list = [
                {
                    'bom': bom_data['bom'],
                    'hours': round(bom_data['hours'], 2),
                    'records': bom_data['records'],
                    'employee_count': len(bom_data['employees'])
                }
                for bom_data in sorted(data['boms'].values(), key=lambda x: x['hours'], reverse=True)
            ]
            
            stats_per_project_list.append({
                'description': project_desc,
                'boms': boms_list,
                'total_hours': round(data['total_hours'], 2),
                'total_records': data['total_records'],
                'employee_count': len(data['employees'])
            })
        
        stats_per_project_list.sort(key=lambda x: x['total_hours'], reverse=True)
        
        all_boms_list = []
        for project_desc, data in stats_per_project.items():
            for bom_id, bom_data in data['boms'].items():
                all_boms_list.append({
                    'bom': bom_id,
                    'project_description': project_desc,
                    'hours': round(bom_data['hours'], 2),
                    'records': bom_data['records'],
                    'employee_count': len(bom_data['employees'])
                })
        all_boms_list.sort(key=lambda x: x['hours'], reverse=True)
        
    except Exception as e:
        logger.error(f"Error calculating project stats: {e}", exc_info=True)
        stats_per_project_list = []
        all_boms_list = []
    
    # Daily trends (current week: Monday to Sunday)
    daily_trends = []
    try:
        # Get Monday of current week
        current_week_start = get_week_start()
        
        # Loop through 7 days from Monday to Sunday
        for i in range(7):
            day_start = current_week_start + timedelta(days=i)
            day_end = day_start + timedelta(days=1)
            day_records = TimeRecord.objects.filter(
                end_time__gte=day_start,
                end_time__lt=day_end
            )
            day_total = sum((r.duration_seconds or 0) for r in day_records) / 3600
            day_prod = sum((r.duration_seconds or 0) for r in day_records if not r.is_non_productive) / 3600
            day_nonprod = sum((r.duration_seconds or 0) for r in day_records if r.is_non_productive) / 3600
            
            # Get day name (Mon, Tue, Wed, etc.)
            day_name = day_start.strftime('%a')
            
            daily_trends.append({
                'date': day_start.strftime('%Y-%m-%d'),
                'day_name': day_name,
                'total_hours': round(day_total, 2),
                'productive_hours': round(day_prod, 2),
                'non_productive_hours': round(day_nonprod, 2),
            })
    except Exception as e:
        logger.error(f"Error calculating daily trends: {e}", exc_info=True)
        daily_trends = []
    
    # Task statistics (top tasks)
    task_stats = {}
    task_stats_list = []
    try:
        for record in month_records:
            if record and not record.is_non_productive and record.task:
                task_name = record.task
                if task_name not in task_stats:
                    task_stats[task_name] = {'hours': 0, 'count': 0}
                task_stats[task_name]['hours'] += (record.duration_seconds or 0) / 3600
                task_stats[task_name]['count'] += 1
        
        task_stats_list = [
            {'name': name, 'hours': round(data.get('hours', 0), 2), 'count': data.get('count', 0)}
            for name, data in task_stats.items()
        ]
        task_stats_list.sort(key=lambda x: x.get('hours', 0), reverse=True)
        task_stats_list = task_stats_list[:10]  # Top 10
    except Exception as e:
        logger.error(f"Error calculating task stats: {e}", exc_info=True)
        task_stats_list = []
    
    # Productivity vs Non-productive (month)
    try:
        month_prod = sum((r.duration_seconds or 0) for r in month_records if r and not r.is_non_productive) / 3600
        month_nonprod = sum((r.duration_seconds or 0) for r in month_records if r and r.is_non_productive) / 3600
    except Exception as e:
        logger.error(f"Error calculating productivity stats: {e}", exc_info=True)
        month_prod = 0
        month_nonprod = 0
    
    # Alerts for long running timers (> 4 hours)
    alerts = []
    try:
        for timer in active_timers:
            if timer and hasattr(timer, 'start_time') and timer.start_time:
                elapsed = (timezone.now() - timer.start_time).total_seconds()
                if elapsed > 4 * 3600:
                    alerts.append({
                        'type': 'long_running',
                        'employee_name': getattr(timer, 'employee_name', 'Unknown'),
                        'task': getattr(timer, 'task', 'Unknown'),
                        'hours': round(elapsed / 3600, 1),
                        'message': f"{getattr(timer, 'employee_name', 'Unknown')} pracuje už {round(elapsed/3600, 1)} hodin"
                    })
    except Exception as e:
        logger.error(f"Error calculating alerts: {e}", exc_info=True)
        alerts = []
    
    # Ensure all lists are initialized
    if not stats_per_person:
        stats_per_person = []
    if not stats_per_project_list:
        stats_per_project_list = []
    if not daily_trends:
        daily_trends = []
    if not task_stats_list:
        task_stats_list = []
    if not alerts:
        alerts = []
    
    return render(request, 'timesheet/admin_dashboard.html', {
        'employees': employee_stats or [],
        'summary': {
            'total_employees': len(employees) if employees else 0,
            'working_now': working_count,
            'on_break': 0,
            'today_total_seconds': total_today,
            'week_total_seconds': total_week,
            'today_total_hours': round(total_today / 3600, 2) if total_today else 0,
            'week_total_hours': round(total_week / 3600, 2) if total_week else 0,
            'month_total_hours': round((month_prod + month_nonprod), 2),
            'month_productive_hours': round(month_prod, 2),
            'month_non_productive_hours': round(month_nonprod, 2),
        },
        'stats_per_person': stats_per_person,
        'stats_per_project': stats_per_project_list,
        'all_boms': all_boms_list,
        'daily_trends': daily_trends,
        'task_stats': task_stats_list,
        'alerts': alerts,
    })


def parse_excel_datetime(date_str, time_str):
    """
    Parse date and time strings from Excel and convert to timezone-aware datetime.
    Excel stores dates as YYYY-MM-DD and times as HH:MM:SS in Excel timezone.
    Also handles HH:MM format from HTML time inputs.
    """
    try:
        # Parse date and time
        if isinstance(date_str, datetime):
            # If Excel already parsed it as datetime
            dt = date_str
        else:
            date_part = str(date_str).strip()
            time_part = str(time_str).strip() if time_str else "00:00:00"
            
            # Combine date and time
            dt_str = f"{date_part} {time_part}"
            
            # Try parsing with seconds first (HH:MM:SS), then without (HH:MM)
            try:
                dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M:%S")
            except ValueError:
                # If that fails, try without seconds (HH:MM format from HTML time inputs)
                dt = datetime.strptime(dt_str, "%Y-%m-%d %H:%M")
        
        # Make timezone-aware using Excel timezone
        excel_tz = get_excel_timezone()
        if timezone.is_naive(dt):
            dt = excel_tz.localize(dt)
        
        # Convert to Django timezone
        django_tz = pytz.timezone(settings.TIME_ZONE)
        dt = dt.astimezone(django_tz)
        
        return dt
    except Exception as e:
        logger.error(f"Error parsing Excel datetime '{date_str} {time_str}': {e}")
        return None


def create_record_key(employee_id, start_time, task, is_non_productive):
    """Create a unique key for matching records"""
    # Use employee_id, start_time (rounded to minute), task, and type
    start_minute = start_time.replace(second=0, microsecond=0)
    return f"{employee_id}|{start_minute.isoformat()}|{task}|{is_non_productive}"


def sync_timesheet_data():
    """
    Sync function: reads Excel, upserts to DB, then replaces Excel with all DB data.
    This ensures Excel always matches the database exactly.
    
    Process:
    1. Read all records from Excel
    2. Upsert each Excel record to database (update if exists, insert if not)
    3. Clear Excel sheets
    4. Write all database records back to Excel
    
    Returns:
        dict: {
            'success': bool,
            'productive_count': int,
            'non_productive_count': int,
            'upserted_from_excel': int,
            'message': str,
            'error': str (if failed)
        }
    """
    excel_client = get_excel_client()
    if not excel_client:
        return {
            'success': False,
            'error': 'Excel client not available. Please set EXCEL_FILE_PATH in environment variables.'
        }
    
    try:
        # Step 1: Read all records from Excel and upsert to database
        upserted_from_excel = 0
        updated_count = 0
        inserted_count = 0
        added_to_excel = 0
        
        # Read productive records from Excel and upsert to database
        try:
            excel_productive = excel_client.get_worksheet_data("Záznamy")
            for row in excel_productive:
                try:
                    employee_id = str(row.get('Zaměstnanec ID', '')).strip()
                    if not employee_id:
                        continue
                    
                    date_str = row.get('Datum')
                    start_time_str = row.get('Začátek')
                    end_time_str = row.get('Konec')
                    
                    if not date_str or not start_time_str or not end_time_str:
                        continue
                    
                    start_dt = parse_excel_datetime(date_str, start_time_str)
                    end_dt = parse_excel_datetime(date_str, end_time_str)
                    
                    if not start_dt or not end_dt:
                        continue
                    
                    # If end time is before start time, assume it's next day
                    if end_dt < start_dt:
                        end_dt = end_dt + timedelta(days=1)
                    
                    task = str(row.get('Úkon', '')).strip()
                    employee_name = str(row.get('Zaměstnanec', '')).strip()
                    project_id = str(row.get('Projekt ID', '')).strip() or None
                    project_name = str(row.get('Projekt', '')).strip() or None
                    
                    # Try to get duration from hours first, then fall back to seconds for backward compatibility
                    duration_hours = row.get('Doba (hodiny)')
                    duration_seconds = row.get('Doba (sekundy)')
                    
                    if duration_hours is not None:
                        # Convert hours to seconds
                        duration_seconds = int(float(duration_hours) * 3600)
                    elif duration_seconds is None:
                        # Calculate from start/end times
                        duration_seconds = int((end_dt - start_dt).total_seconds())
                    else:
                        # Convert to int if it's a float
                        duration_seconds = int(float(duration_seconds))
                    
                    # Upsert to database using composite key (employee_id, start_time rounded to minute, task, is_non_productive)
                    # Round start_time to minute for matching - use a 1-minute window
                    start_minute = start_dt.replace(second=0, microsecond=0)
                    start_window_start = start_minute
                    start_window_end = start_minute + timedelta(minutes=1)
                    
                    # Query for records matching the composite key within the minute window
                    matching_records = TimeRecord.objects.filter(
                        employee_id=employee_id,
                        start_time__gte=start_window_start,
                        start_time__lt=start_window_end,
                        task=task,
                        is_non_productive=False
                    )
                    
                    if matching_records.exists():
                        # Update the first matching record (if duplicates exist, update first one)
                        time_record = matching_records.first()
                        time_record.employee_name = employee_name
                        time_record.project_id = project_id
                        time_record.project_name = project_name
                        time_record.start_time = start_dt
                        time_record.end_time = end_dt
                        time_record.duration_seconds = duration_seconds
                        time_record.save()
                        created = False
                        logger.debug(f"Updated existing productive record: {employee_name} - {task} at {start_dt}")
                    else:
                        # No matching record found, create new one
                        time_record = TimeRecord(
                            id=str(uuid.uuid4()),
                            employee_id=employee_id,
                            employee_name=employee_name,
                            project_id=project_id,
                            project_name=project_name,
                            task=task,
                            is_non_productive=False,
                            start_time=start_dt,
                            end_time=end_dt,
                            duration_seconds=duration_seconds
                        )
                        time_record.save()
                        created = True
                        logger.debug(f"Created new productive record: {employee_name} - {task} at {start_dt}")
                    
                    if created:
                        inserted_count += 1
                        logger.debug(f"Inserted record from Excel to DB: {employee_name} - {task}")
                    else:
                        updated_count += 1
                        logger.debug(f"Updated record from Excel to DB: {employee_name} - {task}")
                    
                    upserted_from_excel += 1
                except Exception as e:
                    logger.warning(f"Error processing Excel productive record: {e}")
                    continue
        except Exception as e:
            logger.warning(f"Error reading productive records from Excel: {e}")
        
        # Read non-productive records from Excel and upsert to database
        try:
            excel_non_productive = excel_client.get_worksheet_data("Neproduktivní záznamy")
            for row in excel_non_productive:
                try:
                    employee_id = str(row.get('Zaměstnanec ID', '')).strip()
                    if not employee_id:
                        continue
                    
                    date_str = row.get('Datum')
                    start_time_str = row.get('Začátek')
                    end_time_str = row.get('Konec')
                    
                    if not date_str or not start_time_str or not end_time_str:
                        continue
                    
                    start_dt = parse_excel_datetime(date_str, start_time_str)
                    end_dt = parse_excel_datetime(date_str, end_time_str)
                    
                    if not start_dt or not end_dt:
                        continue
                    
                    # If end time is before start time, assume it's next day
                    if end_dt < start_dt:
                        end_dt = end_dt + timedelta(days=1)
                    
                    task = str(row.get('Úkon', '')).strip()
                    employee_name = str(row.get('Zaměstnanec', '')).strip()
                    
                    # Try to get duration from hours first, then fall back to seconds for backward compatibility
                    duration_hours = row.get('Doba (hodiny)')
                    duration_seconds = row.get('Doba (sekundy)')
                    
                    if duration_hours is not None:
                        # Convert hours to seconds
                        duration_seconds = int(float(duration_hours) * 3600)
                    elif duration_seconds is None:
                        # Calculate from start/end times
                        duration_seconds = int((end_dt - start_dt).total_seconds())
                    else:
                        # Convert to int if it's a float
                        duration_seconds = int(float(duration_seconds))
                    
                    # Upsert to database using composite key
                    start_minute = start_dt.replace(second=0, microsecond=0)
                    start_window_start = start_minute
                    start_window_end = start_minute + timedelta(minutes=1)
                    
                    # Query for records matching the composite key within the minute window
                    matching_records = TimeRecord.objects.filter(
                        employee_id=employee_id,
                        start_time__gte=start_window_start,
                        start_time__lt=start_window_end,
                        task=task,
                        is_non_productive=True
                    )
                    
                    if matching_records.exists():
                        # Update the first matching record (if duplicates exist, update first one)
                        time_record = matching_records.first()
                        time_record.employee_name = employee_name
                        time_record.project_id = None
                        time_record.project_name = None
                        time_record.start_time = start_dt
                        time_record.end_time = end_dt
                        time_record.duration_seconds = duration_seconds
                        time_record.save()
                        created = False
                        logger.debug(f"Updated existing non-productive record: {employee_name} - {task} at {start_dt}")
                    else:
                        # No matching record found, create new one
                        time_record = TimeRecord(
                            id=str(uuid.uuid4()),
                            employee_id=employee_id,
                            employee_name=employee_name,
                            project_id=None,
                            project_name=None,
                            task=task,
                            is_non_productive=True,
                            start_time=start_dt,
                            end_time=end_dt,
                            duration_seconds=duration_seconds
                        )
                        time_record.save()
                        created = True
                        logger.debug(f"Created new non-productive record: {employee_name} - {task} at {start_dt}")
                    
                    if created:
                        inserted_count += 1
                        logger.debug(f"Inserted non-productive record from Excel to DB: {employee_name} - {task}")
                    else:
                        updated_count += 1
                        logger.debug(f"Updated non-productive record from Excel to DB: {employee_name} - {task}")
                    
                    upserted_from_excel += 1
                except Exception as e:
                    logger.warning(f"Error processing Excel non-productive record: {e}")
                    continue
        except Exception as e:
            logger.warning(f"Error reading non-productive records from Excel: {e}")
        
        logger.info(f"Upserted {upserted_from_excel} records from Excel to DB (inserted: {inserted_count}, updated: {updated_count})")
        
        # Step 2: Get all records from database
        all_db_records = TimeRecord.objects.filter(end_time__isnull=False).order_by('start_time')
        
        # Step 3: Clear Excel sheets and write all database records
        productive_records = []
        non_productive_records = []
        
        for record in all_db_records:
            # Convert to Excel timezone for display
            start_dt_excel = convert_to_excel_timezone(record.start_time)
            end_dt_excel = convert_to_excel_timezone(record.end_time)
            
            # Format duration
            duration_seconds = record.duration_seconds or 0
            hours = duration_seconds // 3600
            minutes = (duration_seconds % 3600) // 60
            seconds = duration_seconds % 60
            duration_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Convert seconds to hours (decimal, rounded to 2 decimal places)
            duration_hours = round(duration_seconds / 3600.0, 2)
            
            if record.is_non_productive:
                row_data = [
                    start_dt_excel.strftime("%Y-%m-%d"),
                    record.employee_id,
                    record.employee_name,
                    record.task,
                    start_dt_excel.strftime("%H:%M:%S"),
                    end_dt_excel.strftime("%H:%M:%S"),
                    duration_formatted,
                    duration_hours
                ]
                non_productive_records.append(row_data)
            else:
                # Productive records have project info
                productive_row = [
                    start_dt_excel.strftime("%Y-%m-%d"),
                    record.employee_id,
                    record.employee_name,
                    record.project_id or '',
                    record.project_name or '',
                    record.task,
                    start_dt_excel.strftime("%H:%M:%S"),
                    end_dt_excel.strftime("%H:%M:%S"),
                    duration_formatted,
                    duration_hours
                ]
                productive_records.append(productive_row)
        
        # Replace Excel worksheets with all database records
        excel_client.replace_worksheet_data(
            "Neproduktivní záznamy",
            ["Datum", "Zaměstnanec ID", "Zaměstnanec", "Úkon",
             "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"],
            non_productive_records
        )
        logger.info(f"Replaced Excel with {len(non_productive_records)} non-productive records")
        
        excel_client.replace_worksheet_data(
            "Záznamy",
            ["Datum", "Zaměstnanec ID", "Zaměstnanec", "Projekt ID", "Projekt",
             "Úkon", "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"],
            productive_records
        )
        logger.info(f"Replaced Excel with {len(productive_records)} productive records")
        
        total_productive = len(productive_records)
        total_non_productive = len(non_productive_records)
        
        return {
            'success': True,
            'message': f'Úspěšně synchronizováno: {total_productive} produktivních záznamů, {total_non_productive} neproduktivních záznamů v DB. Upsertováno z Excel do DB: {upserted_from_excel} (vloženo: {inserted_count}, aktualizováno: {updated_count})',
            'productive_count': total_productive,
            'non_productive_count': total_non_productive,
            'upserted_from_excel': upserted_from_excel,
            'inserted_count': inserted_count,
            'updated_count': updated_count,
        }
        
    except Exception as e:
        logger.error(f"Error syncing to Excel: {e}", exc_info=True)
        return {
            'success': False,
            'error': f'Chyba při synchronizaci: {str(e)}'
        }


# Export Database to Excel (download)
@admin_required
def export_to_excel(request):
    """
    Export all data from database to Excel file and return as download.
    Creates a new Excel file with all worksheets populated from database.
    """
    try:
        # Create a new workbook in memory
        wb = Workbook()
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Get all data from database
        all_records = TimeRecord.objects.filter(end_time__isnull=False).order_by('start_time')
        employees = Employee.objects.filter(is_active=True).order_by('name')
        projects = Project.objects.filter(is_active=True).order_by('name')
        productive_tasks = Task.objects.filter(is_active=True, is_non_productive=False).order_by('name')
        non_productive_tasks = Task.objects.filter(is_active=True, is_non_productive=True).order_by('name')
        
        # Prepare time records
        productive_records = []
        non_productive_records = []
        
        for record in all_records:
            # Convert to Excel timezone for display
            start_dt_excel = convert_to_excel_timezone(record.start_time)
            end_dt_excel = convert_to_excel_timezone(record.end_time)
            
            # Format duration
            duration_seconds = record.duration_seconds or 0
            hours = duration_seconds // 3600
            minutes = (duration_seconds % 3600) // 60
            seconds = duration_seconds % 60
            duration_formatted = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
            
            # Convert seconds to hours (decimal, rounded to 2 decimal places)
            duration_hours = round(duration_seconds / 3600.0, 2)
            
            if record.is_non_productive:
                row_data = [
                    start_dt_excel.strftime("%Y-%m-%d"),
                    record.employee_id,
                    record.employee_name,
                    record.task,
                    start_dt_excel.strftime("%H:%M:%S"),
                    end_dt_excel.strftime("%H:%M:%S"),
                    duration_formatted,
                    duration_hours
                ]
                non_productive_records.append(row_data)
            else:
                productive_row = [
                    start_dt_excel.strftime("%Y-%m-%d"),
                    record.employee_id,
                    record.employee_name,
                    record.project_id or '',
                    record.project_name or '',
                    record.task,
                    start_dt_excel.strftime("%H:%M:%S"),
                    end_dt_excel.strftime("%H:%M:%S"),
                    duration_formatted,
                    duration_hours
                ]
                productive_records.append(productive_row)
        
        # Create worksheets and populate data
        
        # 1. Productive records (Záznamy)
        ws_prod = wb.create_sheet("Záznamy")
        ws_prod.append(["Datum", "Zaměstnanec ID", "Zaměstnanec", "Projekt ID", "Projekt",
                       "Úkon", "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"])
        for row in productive_records:
            ws_prod.append(row)
        
        # 2. Non-productive records (Neproduktivní záznamy)
        ws_nonprod = wb.create_sheet("Neproduktivní záznamy")
        ws_nonprod.append(["Datum", "Zaměstnanec ID", "Zaměstnanec", "Úkon",
                          "Začátek", "Konec", "Doba trvání", "Doba (hodiny)"])
        for row in non_productive_records:
            ws_nonprod.append(row)
        
        # 3. Employees (Zaměstnanci)
        ws_emp = wb.create_sheet("Zaměstnanci")
        ws_emp.append(["ID", "Jméno"])
        for emp in employees:
            ws_emp.append([emp.id, emp.name])
        
        # 4. Projects (Projekty)
        ws_proj = wb.create_sheet("Projekty")
        ws_proj.append(["ID", "Název"])
        for proj in projects:
            ws_proj.append([proj.id, proj.name])
        
        # 5. Productive tasks (Úkony)
        ws_tasks = wb.create_sheet("Úkony")
        ws_tasks.append(["Název"])
        for task in productive_tasks:
            ws_tasks.append([task.name])
        
        # 6. Non-productive tasks (Neproduktivní úkony)
        ws_nonprod_tasks = wb.create_sheet("Neproduktivní úkony")
        ws_nonprod_tasks.append(["Název"])
        for task in non_productive_tasks:
            ws_nonprod_tasks.append([task.name])
        
        # Save workbook to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        now = datetime.now()
        filename = f"TimeSheet_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        # Create HTTP response with Excel file
        response = HttpResponse(
            output.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        logger.info(f"Exported Excel file: {filename} ({len(productive_records)} productive, {len(non_productive_records)} non-productive records)")
        
        return response
        
    except Exception as e:
        logger.error(f"Error exporting to Excel: {e}", exc_info=True)
        # Safely encode error message to handle Czech characters
        try:
            error_msg = str(e)
            # Ensure it's a valid UTF-8 string
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při exportu: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})


# Sync Database and Excel (bidirectional)
@login_required
def sync_to_excel(request):
    """
    Bidirectional sync: ensures all records in Excel exist in DB and vice versa.
    Reads from both sources, merges them, and updates both.
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    result = sync_timesheet_data()
    
    if result['success']:
        return JsonResponse(result, json_dumps_params={'ensure_ascii': False})
    else:
        return JsonResponse(result, status=500, json_dumps_params={'ensure_ascii': False})


# Admin Control Panel
@admin_required
def admin_control_panel(request):
    """Admin control panel for managing employees, projects, and tasks"""
    # Get all active records
    employees = Employee.objects.filter(is_active=True).order_by('name')
    projects = Project.objects.filter(is_active=True).order_by('name')
    productive_tasks = Task.objects.filter(is_active=True, is_non_productive=False).order_by('name')
    non_productive_tasks = Task.objects.filter(is_active=True, is_non_productive=True).order_by('name')
    
    # Format for display in textarea (ID\tName format)
    employees_text = '\n'.join([f"{emp.id}\t{emp.name}" for emp in employees])
    projects_text = '\n'.join([f"{proj.id}\t{proj.name} - {proj.project_description}" if proj.project_description else f"{proj.id}\t{proj.name}" for proj in projects])
    productive_tasks_text = '\n'.join([task.name for task in productive_tasks])
    non_productive_tasks_text = '\n'.join([task.name for task in non_productive_tasks])
    
    return render(request, 'timesheet/admin_control_panel.html', {
        'employees_text': employees_text,
        'projects_text': projects_text,
        'productive_tasks_text': productive_tasks_text,
        'non_productive_tasks_text': non_productive_tasks_text,
        'employees_count': employees.count(),
        'projects_count': projects.count(),
        'productive_tasks_count': productive_tasks.count(),
        'non_productive_tasks_count': non_productive_tasks.count(),
    })


@admin_required
@transaction.atomic
def save_master_data(request):
    """Save master data from admin control panel"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        employees_text = request.POST.get('employees', '').strip()
        projects_text = request.POST.get('projects', '').strip()
        productive_tasks_text = request.POST.get('productive_tasks', '').strip()
        non_productive_tasks_text = request.POST.get('non_productive_tasks', '').strip()
        
        if isinstance(employees_text, bytes):
            employees_text = employees_text.decode('utf-8', errors='replace')
        if isinstance(projects_text, bytes):
            projects_text = projects_text.decode('utf-8', errors='replace')
        if isinstance(productive_tasks_text, bytes):
            productive_tasks_text = productive_tasks_text.decode('utf-8', errors='replace')
        if isinstance(non_productive_tasks_text, bytes):
            non_productive_tasks_text = non_productive_tasks_text.decode('utf-8', errors='replace')
        
        logger.info(f"Received non-productive tasks text length: {len(non_productive_tasks_text)}")
        if 'VYKLÁDKA' in non_productive_tasks_text:
            logger.info("Found VYKLÁDKA in non-productive tasks text")
        
        productive_task_names = set()
        if productive_tasks_text:
            for line in productive_tasks_text.splitlines():
                task_name = line.strip()
                if task_name:
                    try:
                        if isinstance(task_name, bytes):
                            task_name = task_name.decode('utf-8', errors='replace')
                        if not isinstance(task_name, str):
                            task_name = str(task_name)
                        task_name = task_name.strip()
                        if task_name:
                            productive_task_names.add(task_name)
                    except Exception as e:
                        logger.error(f"Error processing productive task '{line}': {e}", exc_info=True)
        
        # Parse and save employees (format: ID\tName)
        if employees_text:
            existing_employee_ids = set(Employee.objects.values_list('id', flat=True))
            new_employee_ids = set()
            
            # Use splitlines() to handle all line ending types (\n, \r\n, \r)
            for line in employees_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                
                parts = line.split('\t')
                if len(parts) >= 2:
                    emp_id = parts[0].strip()
                    emp_name = '\t'.join(parts[1:]).strip()
                else:
                    # Try space separation
                    parts = line.split(None, 1)
                    if len(parts) == 2:
                        emp_id = parts[0].strip()
                        emp_name = parts[1].strip()
                    else:
                        continue
                
                if emp_id and emp_name:
                    new_employee_ids.add(emp_id)
                    Employee.objects.update_or_create(
                        id=emp_id,
                        defaults={'name': emp_name, 'is_active': True}
                    )
            
            # Deactivate employees not in the new list (only if we have valid employees)
            if new_employee_ids:
                Employee.objects.exclude(id__in=new_employee_ids).update(is_active=False)
        
        # Parse and save projects (format: ID\tName)
        if projects_text:
            existing_project_ids = set(Project.objects.values_list('id', flat=True))
            new_project_ids = set()
            
            # Use splitlines() to handle all line ending types (\n, \r\n, \r)
            for line in projects_text.splitlines():
                line = line.strip()
                if not line:
                    continue
                
                parts = line.split('\t')
                if len(parts) >= 2:
                    proj_id = parts[0].strip()
                    proj_name = '\t'.join(parts[1:]).strip()
                else:
                    parts = line.split(None, 1)
                    if len(parts) == 2:
                        proj_id = parts[0].strip()
                        proj_name = parts[1].strip()
                    else:
                        continue
                
                if proj_id and proj_name:
                    new_project_ids.add(proj_id)
                    
                    project_description = None
                    if ' - ' in proj_name:
                        name_parts = proj_name.split(' - ', 1)
                        proj_name_clean = name_parts[0].strip()
                        project_description = name_parts[1].strip() if len(name_parts) > 1 else None
                    else:
                        proj_name_clean = proj_name
                    
                    Project.objects.update_or_create(
                        id=proj_id,
                        defaults={
                            'name': proj_id,
                            'project_description': project_description,
                            'is_active': True
                        }
                    )
            
            # Deactivate projects not in the new list (only if we have valid projects)
            if new_project_ids:
                Project.objects.exclude(id__in=new_project_ids).update(is_active=False)
        
        # Parse and save productive tasks (format: Name per line)
        if productive_tasks_text:
            new_task_names = set()
            
            # Use splitlines() to handle all line ending types (\n, \r\n, \r)
            lines = productive_tasks_text.splitlines()
            
            for line in lines:
                task_name = line.strip()
                if not task_name:
                    continue
                
                try:
                    if isinstance(task_name, bytes):
                        task_name = task_name.decode('utf-8', errors='replace')
                    if not isinstance(task_name, str):
                        task_name = str(task_name)
                except (UnicodeEncodeError, UnicodeDecodeError) as e:
                    logger.error(f"Encoding error for task name '{line}': {e}", exc_info=True)
                    continue
                
                if not task_name:
                    continue
                
                new_task_names.add(task_name)
                try:
                    existing_task = Task.objects.filter(name=task_name, is_non_productive=False).first()
                    if not existing_task:
                        existing_task = Task.objects.filter(name__iexact=task_name, is_non_productive=False).first()
                    
                    if existing_task:
                        existing_task.is_non_productive = False
                        existing_task.is_active = True
                        existing_task.save()
                        new_task_names.discard(task_name)
                        new_task_names.add(existing_task.name)
                    else:
                        Task.objects.create(
                            name=task_name,
                            is_non_productive=False,
                            is_active=True
                        )
                except Exception as e:
                    logger.error(f"Error saving task '{task_name}': {e}", exc_info=True)
                    continue
            
            # Deactivate productive tasks not in the new list (only if we have valid tasks)
            if new_task_names:
                Task.objects.filter(is_non_productive=False).exclude(name__in=new_task_names).update(is_active=False)
        
        # Parse and save non-productive tasks (format: Name per line)
        if non_productive_tasks_text:
            new_task_names = set()
            
            # Use splitlines() to handle all line ending types (\n, \r\n, \r)
            lines = non_productive_tasks_text.splitlines()
            
            for line in lines:
                task_name = line.strip()
                if not task_name:
                    continue
                
                try:
                    if isinstance(task_name, bytes):
                        task_name = task_name.decode('utf-8', errors='replace')
                    if not isinstance(task_name, str):
                        task_name = str(task_name)
                    task_name = task_name.strip()
                except Exception as e:
                    logger.error(f"Encoding error for task name '{repr(line)}': {e}", exc_info=True)
                    continue
                
                if not task_name:
                    continue
                
                new_task_names.add(task_name)
                try:
                    existing_task = Task.objects.filter(name=task_name, is_non_productive=True).first()
                    if not existing_task:
                        existing_task = Task.objects.filter(name__iexact=task_name, is_non_productive=True).first()
                    
                    if existing_task:
                        existing_task.is_non_productive = True
                        existing_task.is_active = True
                        existing_task.save()
                        new_task_names.discard(task_name)
                        new_task_names.add(existing_task.name)
                    else:
                        try:
                            new_task = Task(
                                name=task_name,
                                is_non_productive=True,
                                is_active=True
                            )
                            new_task.full_clean()
                            new_task.save()
                            logger.info(f"Successfully created non-productive task: '{task_name}' (length: {len(task_name)})")
                        except Exception as create_error:
                            logger.error(f"Failed to create task '{task_name}': {create_error}", exc_info=True)
                            raise
                except Exception as e:
                    error_detail = str(e)
                    logger.error(f"Error saving task '{task_name}': {error_detail}", exc_info=True)
                    continue
            
            # Deactivate non-productive tasks not in the new list (only if we have valid tasks)
            if new_task_names:
                Task.objects.filter(is_non_productive=True).exclude(name__in=new_task_names).update(is_active=False)
        
        return JsonResponse({
            'success': True,
            'message': 'Data byla úspěšně uložena'
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        logger.error(f"Error saving master data: {e}", exc_info=True)
        import traceback
        error_trace = traceback.format_exc()
        logger.error(f"Full traceback: {error_trace}")
        
        try:
            error_msg = str(e)
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při ukládání: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})


# Edit Time Records Page
@login_required
def edit_time_records(request):
    """Page for editing time records with filtering and inline editing"""
    # Get filter parameters
    employee_filter = request.GET.get('employee', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    is_non_productive_filter = request.GET.get('is_non_productive', '')
    
    # Set default date_from to beginning of current week (Monday) if not provided
    if not date_from:
        today = timezone.now().date()
        # Get Monday of current week (weekday() returns 0=Monday, 6=Sunday)
        days_since_monday = today.weekday()
        monday = today - timedelta(days=days_since_monday)
        date_from = monday.strftime('%Y-%m-%d')
    
    # Set default date_to to today if not provided
    if not date_to:
        today = timezone.now().date()
        date_to = today.strftime('%Y-%m-%d')
    
    # Build query
    records_query = TimeRecord.objects.filter(end_time__isnull=False)
    
    if employee_filter:
        records_query = records_query.filter(employee_id=employee_filter)
    
    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
            records_query = records_query.filter(start_time__date__gte=from_date)
        except ValueError:
            pass
    
    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
            records_query = records_query.filter(start_time__date__lte=to_date)
        except ValueError:
            pass
    
    if is_non_productive_filter == 'true':
        records_query = records_query.filter(is_non_productive=True)
    elif is_non_productive_filter == 'false':
        records_query = records_query.filter(is_non_productive=False)
    
    # Order by date (newest first)
    records = records_query.order_by('-start_time')[:1000]  # Limit to 1000 records
    
    # Format records for display
    records_data = []
    for record in records:
        start_dt_excel = convert_to_excel_timezone(record.start_time)
        end_dt_excel = convert_to_excel_timezone(record.end_time) if record.end_time else None
        
        # Calculate hours from start_time and end_time
        calculated_hours = None
        if record.start_time and record.end_time:
            duration_delta = record.end_time - record.start_time
            calculated_seconds = int(duration_delta.total_seconds())
            calculated_hours = round(calculated_seconds / 3600.0, 2)
        
        # Use stored duration_hours from database (duration_seconds converted to hours)
        # Prefer stored value if it exists and is > 0, otherwise use calculated
        duration_seconds = record.duration_seconds or 0
        if duration_seconds and duration_seconds > 0:
            duration_hours = round(duration_seconds / 3600.0, 2)
        elif calculated_hours is not None:
            # Use calculated hours if stored value is missing/zero
            duration_hours = calculated_hours
        else:
            # Fallback to 0
            duration_hours = 0.0
        
        # Ensure duration_hours is always a float
        duration_hours = float(duration_hours) if duration_hours is not None else 0.0
        
        # Format calculated_hours
        if calculated_hours is not None:
            try:
                calculated_hours = float(calculated_hours)
            except (TypeError, ValueError):
                calculated_hours = None
        
        record_dict = {
            'id': record.id,
            'date': start_dt_excel.strftime('%Y-%m-%d'),
            'employee_id': record.employee_id,
            'employee_name': record.employee_name,
            'project_id': record.project_id or '',
            'project_name': record.project_name or '',
            'task': record.task,
            'start_time': start_dt_excel.strftime('%H:%M:%S'),
            'end_time': end_dt_excel.strftime('%H:%M:%S') if end_dt_excel else '',
            'duration_hours': round(duration_hours, 2),  # Ensure 2 decimal places
            'calculated_hours': round(calculated_hours, 2) if calculated_hours is not None else None,
            'duration_seconds': duration_seconds,
            'is_non_productive': record.is_non_productive,
        }
        
        records_data.append(record_dict)
    
    # Calculate totals
    total_hours = sum(rec['duration_hours'] for rec in records_data)
    total_productive_hours = sum(rec['duration_hours'] for rec in records_data if not rec['is_non_productive'])
    total_non_productive_hours = sum(rec['duration_hours'] for rec in records_data if rec['is_non_productive'])
    total_records = len(records_data)
    
    # Get employees, projects, and tasks for dropdowns
    employees = Employee.objects.filter(is_active=True).order_by('name')
    projects = Project.objects.filter(is_active=True).order_by('name')
    productive_tasks = Task.objects.filter(is_active=True, is_non_productive=False).order_by('name')
    non_productive_tasks = Task.objects.filter(is_active=True, is_non_productive=True).order_by('name')
    
    return render(request, 'timesheet/edit_time_records.html', {
        'records': records_data,
        'employees': employees,
        'projects': projects,
        'productive_tasks': productive_tasks,
        'non_productive_tasks': non_productive_tasks,
        'filters': {
            'employee': employee_filter,
            'date_from': date_from,
            'date_to': date_to,
            'is_non_productive': is_non_productive_filter,
        },
        'summary': {
            'total_hours': round(total_hours, 2),
            'total_productive_hours': round(total_productive_hours, 2),
            'total_non_productive_hours': round(total_non_productive_hours, 2),
            'total_records': total_records,
        },
    })


@admin_required
def save_time_record(request):
    """Save updated time record - handles all editable fields"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        record_id = request.POST.get('id')
        
        if not record_id:
            return JsonResponse({
                'success': False,
                'error': 'Missing record ID'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Get record
        try:
            record = TimeRecord.objects.get(id=record_id)
        except TimeRecord.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Record not found'
            }, status=404, json_dumps_params={'ensure_ascii': False})
        
        # Update fields if provided
        updated_fields = []
        
        # Update date and/or start time
        date_str = request.POST.get('date')
        start_time_str = request.POST.get('start_time')
        
        # If date or start_time is provided, update start_time
        # (both are sent together from frontend for proper parsing)
        if date_str and start_time_str:
            try:
                start_dt = parse_excel_datetime(date_str, start_time_str)
                if start_dt:
                    record.start_time = start_dt
                    updated_fields.append('start_time')
            except Exception as e:
                logger.warning(f"Error parsing start time: {e}")
        
        # Update end time
        if request.POST.get('end_time'):
            try:
                date_str = request.POST.get('date') or record.start_time.strftime('%Y-%m-%d')
                time_str = request.POST.get('end_time')
                end_dt = parse_excel_datetime(date_str, time_str)
                if end_dt:
                    # If end time is before start time, assume next day
                    if end_dt < record.start_time:
                        end_dt = end_dt + timedelta(days=1)
                    record.end_time = end_dt
                    updated_fields.append('end_time')
            except Exception as e:
                logger.warning(f"Error parsing end time: {e}")
        
        # Update project
        if 'project_id' in request.POST:
            project_id = request.POST.get('project_id', '').strip()
            if project_id:
                try:
                    project = Project.objects.get(id=project_id, is_active=True)
                    record.project_id = project.id
                    record.project_name = project.project_description or project.name
                    updated_fields.append('project')
                except Project.DoesNotExist:
                    # Clear project if not found
                    record.project_id = None
                    record.project_name = None
                    updated_fields.append('project')
            else:
                # Empty project ID means clear it
                record.project_id = None
                record.project_name = None
                updated_fields.append('project')
        
        # Update task
        if 'task' in request.POST:
            task_name = request.POST.get('task', '').strip()
            if task_name:
                record.task = task_name
                updated_fields.append('task')
        
        # Update duration (hours) - this will recalculate end_time if not explicitly set
        if request.POST.get('duration_hours'):
            try:
                duration_hours = float(request.POST.get('duration_hours'))
                duration_seconds = int(duration_hours * 3600)
                record.duration_seconds = duration_seconds
                updated_fields.append('duration')
                
                # Recalculate end_time based on duration if end_time wasn't explicitly updated
                if 'end_time' not in updated_fields and record.start_time:
                    record.end_time = record.start_time + timedelta(seconds=duration_seconds)
            except (ValueError, TypeError) as e:
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid duration value'
                }, status=400, json_dumps_params={'ensure_ascii': False})
        elif 'end_time' in updated_fields and record.start_time and record.end_time:
            # Recalculate duration from start and end times if duration wasn't provided
            duration_seconds = int((record.end_time - record.start_time).total_seconds())
            record.duration_seconds = duration_seconds
            updated_fields.append('duration')
        
        record.save()
        
        logger.info(f"Updated time record {record_id}: {', '.join(updated_fields)}")
        
        # Format response data
        start_dt_excel = convert_to_excel_timezone(record.start_time)
        end_dt_excel = convert_to_excel_timezone(record.end_time) if record.end_time else None
        duration_hours = round((record.duration_seconds or 0) / 3600.0, 2)
        
        return JsonResponse({
            'success': True,
            'message': 'Záznam byl úspěšně uložen',
            'date': start_dt_excel.strftime('%Y-%m-%d'),
            'start_time': start_dt_excel.strftime('%H:%M:%S'),
            'end_time': end_dt_excel.strftime('%H:%M:%S') if end_dt_excel else '',
            'duration_hours': duration_hours,
            'project_id': record.project_id or '',
            'project_name': record.project_name or '',
            'task': record.task,
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        logger.error(f"Error saving time record: {e}", exc_info=True)
        try:
            error_msg = str(e)
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při ukládání: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})


@admin_required
def delete_time_record(request):
    """Delete a time record"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        record_id = request.POST.get('id')
        
        if not record_id:
            return JsonResponse({
                'success': False,
                'error': 'Missing record ID'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Get and delete record
        try:
            record = TimeRecord.objects.get(id=record_id)
            record.delete()
            logger.info(f"Deleted time record {record_id}")
            
            return JsonResponse({
                'success': True,
                'message': 'Záznam byl úspěšně smazán'
            }, json_dumps_params={'ensure_ascii': False})
            
        except TimeRecord.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Record not found'
            }, status=404, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        logger.error(f"Error deleting time record: {e}", exc_info=True)
        try:
            error_msg = str(e)
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při mazání: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})


@admin_required
def bulk_save_time_records(request):
    """Bulk save/update/delete time records"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Parse JSON body
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        updates = data.get('updates', [])
        deletes = data.get('deletes', [])
        
        if not updates and not deletes:
            return JsonResponse({
                'success': False,
                'error': 'No changes to save'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        results = {'updated': [], 'deleted': [], 'errors': []}
        
        with transaction.atomic():
            # Process updates
            records_to_update = []
            record_ids_to_update = []
            
            for update_data in updates:
                record_id = update_data.get('id')
                if not record_id:
                    results['errors'].append({'id': None, 'error': 'Missing record ID'})
                    continue
                
                try:
                    record = TimeRecord.objects.get(id=record_id)
                except TimeRecord.DoesNotExist:
                    results['errors'].append({'id': record_id, 'error': 'Record not found'})
                    continue
                
                # Update fields
                updated = False
                
                # Update date and/or start time
                date_str = update_data.get('date')
                start_time_str = update_data.get('start_time')
                if date_str and start_time_str:
                    try:
                        start_dt = parse_excel_datetime(date_str, start_time_str)
                        if start_dt:
                            record.start_time = start_dt
                            updated = True
                    except Exception as e:
                        logger.warning(f"Error parsing start time for record {record_id}: {e}")
                        results['errors'].append({'id': record_id, 'error': f'Invalid start time: {e}'})
                        continue
                
                # Update end time
                if 'end_time' in update_data:
                    try:
                        date_str = update_data.get('date') or (record.start_time.strftime('%Y-%m-%d') if record.start_time else None)
                        time_str = update_data.get('end_time')
                        if date_str and time_str:
                            end_dt = parse_excel_datetime(date_str, time_str)
                            if end_dt:
                                # If end time is before start time, assume next day
                                if record.start_time and end_dt < record.start_time:
                                    end_dt = end_dt + timedelta(days=1)
                                record.end_time = end_dt
                                updated = True
                    except Exception as e:
                        logger.warning(f"Error parsing end time for record {record_id}: {e}")
                        results['errors'].append({'id': record_id, 'error': f'Invalid end time: {e}'})
                        continue
                
                # Update employee
                if 'employee_id' in update_data:
                    employee_id = update_data.get('employee_id', '').strip()
                    if employee_id:
                        try:
                            employee = Employee.objects.get(id=employee_id, is_active=True)
                            record.employee_id = employee.id
                            record.employee_name = employee.name
                            updated = True
                        except Employee.DoesNotExist:
                            results['errors'].append({'id': record_id, 'error': f'Employee not found: {employee_id}'})
                            continue
                    else:
                        results['errors'].append({'id': record_id, 'error': 'Employee ID cannot be empty'})
                        continue
                
                # Update project
                if 'project_id' in update_data:
                    project_id = update_data.get('project_id', '').strip()
                    if project_id:
                        try:
                            project = Project.objects.get(id=project_id, is_active=True)
                            record.project_id = project.id
                            record.project_name = project.project_description or project.name
                            updated = True
                        except Project.DoesNotExist:
                            record.project_id = None
                            record.project_name = None
                            updated = True
                    else:
                        record.project_id = None
                        record.project_name = None
                        updated = True
                
                # Update task
                if 'task' in update_data:
                    task_name = update_data.get('task', '').strip()
                    if task_name:
                        record.task = task_name
                        updated = True
                
                # Update duration (hours)
                if 'duration_hours' in update_data:
                    try:
                        duration_hours = float(update_data.get('duration_hours'))
                        duration_seconds = int(duration_hours * 3600)
                        record.duration_seconds = duration_seconds
                        updated = True
                        
                        # Recalculate end_time based on duration if end_time wasn't explicitly updated
                        if 'end_time' not in update_data and record.start_time:
                            record.end_time = record.start_time + timedelta(seconds=duration_seconds)
                    except (ValueError, TypeError) as e:
                        results['errors'].append({'id': record_id, 'error': f'Invalid duration value: {e}'})
                        continue
                
                # Recalculate duration from start and end times if both are set but duration wasn't provided
                if 'end_time' in update_data and 'duration_hours' not in update_data and record.start_time and record.end_time:
                    duration_seconds = int((record.end_time - record.start_time).total_seconds())
                    record.duration_seconds = duration_seconds
                    updated = True
                
                # If start_time changed and duration_hours is set, recalculate end_time
                if ('start_time' in update_data or 'date' in update_data) and 'duration_hours' in update_data and 'end_time' not in update_data:
                    if record.start_time and record.duration_seconds:
                        record.end_time = record.start_time + timedelta(seconds=record.duration_seconds)
                        updated = True
                
                if updated:
                    records_to_update.append(record)
                    record_ids_to_update.append(record_id)
            
            # Bulk update records
            if records_to_update:
                TimeRecord.objects.bulk_update(
                    records_to_update,
                    ['start_time', 'end_time', 'duration_seconds', 'employee_id', 'employee_name', 'project_id', 'project_name', 'task'],
                    batch_size=100
                )
                results['updated'] = record_ids_to_update
                logger.info(f"Bulk updated {len(records_to_update)} time records")
            
            # Process deletes
            if deletes:
                deleted_count = TimeRecord.objects.filter(id__in=deletes).delete()[0]
                results['deleted'] = deletes
                logger.info(f"Bulk deleted {deleted_count} time records")
        
        # Format response with updated record data
        updated_records_data = []
        if results['updated']:
            updated_records = TimeRecord.objects.filter(id__in=results['updated'])
            for record in updated_records:
                start_dt_excel = convert_to_excel_timezone(record.start_time)
                end_dt_excel = convert_to_excel_timezone(record.end_time) if record.end_time else None
                duration_hours = round((record.duration_seconds or 0) / 3600.0, 2)
                
                updated_records_data.append({
                    'id': str(record.id),
                    'date': start_dt_excel.strftime('%Y-%m-%d'),
                    'start_time': start_dt_excel.strftime('%H:%M:%S'),
                    'end_time': end_dt_excel.strftime('%H:%M:%S') if end_dt_excel else '',
                    'duration_hours': duration_hours,
                    'employee_id': record.employee_id or '',
                    'employee_name': record.employee_name or '',
                    'project_id': record.project_id or '',
                    'project_name': record.project_name or '',
                    'task': record.task,
                })
        
        return JsonResponse({
            'success': True,
            'message': f'Successfully updated {len(results["updated"])} record(s) and deleted {len(results["deleted"])} record(s)',
            'updated': updated_records_data,
            'deleted': results['deleted'],
            'errors': results['errors']
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        logger.error(f"Error in bulk save: {e}", exc_info=True)
        try:
            error_msg = str(e)
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při hromadném ukládání: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})


# Daily top-up constants: only top up days with 4h < total < 6.91h to reach 6.91h
DAILY_MIN_HOURS = 4.0
DAILY_TARGET_HOURS = 6.91


@admin_required
def topup_preview(request):
    """Preview which records would get top-up hours to reach 6.91h per day.
    Applies only to days where employee has more than 4h but less than 6.91h.
    Remaining hours are split by the percentage each activity already has.
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Only GET allowed'}, status=405)
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    employee = request.GET.get('employee', '').strip()

    if not date_from or not date_to:
        return JsonResponse({
            'success': False,
            'error': 'date_from and date_to are required'
        }, status=400, json_dumps_params={'ensure_ascii': False})

    try:
        from_date = datetime.strptime(date_from, '%Y-%m-%d').date()
        to_date = datetime.strptime(date_to, '%Y-%m-%d').date()
    except ValueError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid date format (use YYYY-MM-DD)'
        }, status=400, json_dumps_params={'ensure_ascii': False})

    records_query = TimeRecord.objects.filter(
        end_time__isnull=False,
        start_time__date__gte=from_date,
        start_time__date__lte=to_date,
    )
    if employee:
        records_query = records_query.filter(employee_id=employee)
    records = list(records_query.order_by('employee_id', 'start_time'))

    from collections import defaultdict
    by_key = defaultdict(list)
    for r in records:
        dur_sec = r.duration_seconds or 0
        if r.start_time and r.end_time and (not r.duration_seconds or r.duration_seconds <= 0):
            dur_sec = int((r.end_time - r.start_time).total_seconds())
        hours = dur_sec / 3600.0
        key = (r.employee_id, r.employee_name, r.start_time.date())
        by_key[key].append((r, hours))

    suggestions = []
    for (emp_id, emp_name, d) in sorted(by_key.keys(), key=lambda k: (k[0], k[2])):
        recs = by_key[(emp_id, emp_name, d)]
        total = sum(h for _, h in recs)
        if total <= DAILY_MIN_HOURS or total >= DAILY_TARGET_HOURS:
            continue
        remaining = DAILY_TARGET_HOURS - total
        total_sec = sum(int(round(h * 3600)) for _, h in recs)
        remaining_sec = int(round(remaining * 3600))
        if total_sec <= 0:
            continue
        assigned = 0
        for i, (record, rec_hours) in enumerate(recs):
            rec_sec = int(round(rec_hours * 3600))
            if i == len(recs) - 1:
                add_sec = remaining_sec - assigned
            else:
                add_sec = int(round(remaining_sec * (rec_sec / total_sec)))
            assigned += add_sec
            add_hours = round(add_sec / 3600.0, 2)
            new_hours = round(rec_hours + add_hours, 2)
            suggestions.append({
                'record_id': str(record.id),
                'employee_id': emp_id,
                'employee_name': emp_name,
                'date': d.strftime('%Y-%m-%d'),
                'task': record.task,
                'project_name': (record.project_name or '-').strip() or '-',
                'is_non_productive': record.is_non_productive,
                'current_hours': round(rec_hours, 2),
                'add_hours': add_hours,
                'new_hours': new_hours,
            })

    employees_affected = len(set(s['employee_id'] for s in suggestions))
    days_affected = len(set((s['employee_id'], s['date']) for s in suggestions))

    return JsonResponse({
        'success': True,
        'suggestions': suggestions,
        'summary': {
            'total_records': len(suggestions),
            'employees_affected': employees_affected,
            'days_affected': days_affected,
        },
    }, json_dumps_params={'ensure_ascii': False})


@admin_required
def topup_apply(request):
    """Apply top-up hours to the given records (add_hours added to each)."""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST allowed'}, status=405)

    try:
        data = json.loads(request.body)
    except json.JSONDecodeError:
        return JsonResponse({
            'success': False,
            'error': 'Invalid JSON'
        }, status=400, json_dumps_params={'ensure_ascii': False})

    updates = data.get('updates', [])
    if not updates:
        return JsonResponse({
            'success': False,
            'error': 'No updates provided'
        }, status=400, json_dumps_params={'ensure_ascii': False})

    applied = []
    errors = []
    with transaction.atomic():
        for item in updates:
            rid = item.get('record_id')
            add_hours = item.get('add_hours', 0)
            if not rid:
                errors.append({'record_id': None, 'error': 'Missing record_id'})
                continue
            try:
                add_hours = float(add_hours)
            except (TypeError, ValueError):
                errors.append({'record_id': rid, 'error': 'Invalid add_hours'})
                continue
            try:
                record = TimeRecord.objects.get(id=rid)
            except TimeRecord.DoesNotExist:
                errors.append({'record_id': rid, 'error': 'Record not found'})
                continue
            prev_sec = record.duration_seconds or 0
            if record.start_time and record.end_time and prev_sec <= 0:
                prev_sec = int((record.end_time - record.start_time).total_seconds())
            add_sec = int(round(add_hours * 3600))
            new_sec = prev_sec + add_sec
            record.duration_seconds = new_sec
            record.end_time = record.start_time + timedelta(seconds=new_sec)
            record.save()
            applied.append(rid)

    return JsonResponse({
        'success': True,
        'message': f'Updated {len(applied)} record(s).',
        'updated': applied,
        'errors': errors,
    }, json_dumps_params={'ensure_ascii': False})


@admin_required
def create_time_record(request):
    """Create a new time record"""
    if request.method != 'POST':
        return JsonResponse({'error': 'Only POST method allowed'}, status=405)
    
    try:
        # Parse JSON body
        try:
            data = json.loads(request.body)
        except json.JSONDecodeError:
            return JsonResponse({
                'success': False,
                'error': 'Invalid JSON data'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Validate required fields
        required_fields = ['date', 'employee_id', 'task', 'start_time', 'is_non_productive']
        for field in required_fields:
            if field not in data:
                return JsonResponse({
                    'success': False,
                    'error': f'Missing required field: {field}'
                }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Get employee
        try:
            employee = Employee.objects.get(id=data['employee_id'], is_active=True)
        except Employee.DoesNotExist:
            return JsonResponse({
                'success': False,
                'error': 'Employee not found'
            }, status=404, json_dumps_params={'ensure_ascii': False})
        
        # Parse start time
        try:
            start_dt = parse_excel_datetime(data['date'], data['start_time'])
            if not start_dt:
                raise ValueError('Invalid date/time format')
        except Exception as e:
            return JsonResponse({
                'success': False,
                'error': f'Invalid start time: {e}'
            }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Parse end time if provided
        end_dt = None
        if data.get('end_time'):
            try:
                end_dt = parse_excel_datetime(data['date'], data['end_time'])
                if end_dt and end_dt < start_dt:
                    end_dt = end_dt + timedelta(days=1)
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'error': f'Invalid end time: {e}'
                }, status=400, json_dumps_params={'ensure_ascii': False})
        
        # Calculate duration
        duration_seconds = None
        if data.get('duration_hours'):
            try:
                duration_hours = float(data['duration_hours'])
                duration_seconds = int(duration_hours * 3600)
            except (ValueError, TypeError):
                return JsonResponse({
                    'success': False,
                    'error': 'Invalid duration value'
                }, status=400, json_dumps_params={'ensure_ascii': False})
        elif end_dt:
            # Calculate from start and end times
            duration_seconds = int((end_dt - start_dt).total_seconds())
        
        # If we have duration but no end_time, calculate end_time
        if duration_seconds and not end_dt:
            end_dt = start_dt + timedelta(seconds=duration_seconds)
        
        # Get project if provided
        project_id = None
        project_name = None
        if data.get('project_id'):
            try:
                project = Project.objects.get(id=data['project_id'], is_active=True)
                project_id = project.id
                project_name = project.project_description or project.name
            except Project.DoesNotExist:
                pass
        
        # Create record
        record = TimeRecord.objects.create(
            employee_id=employee.id,
            employee_name=employee.name,
            project_id=project_id,
            project_name=project_name,
            task=data['task'],
            is_non_productive=data['is_non_productive'],
            start_time=start_dt,
            end_time=end_dt,
            duration_seconds=duration_seconds,
        )
        
        logger.info(f"Created new time record {record.id} for employee {employee.name}")
        
        return JsonResponse({
            'success': True,
            'message': 'Record created successfully',
            'id': str(record.id),
        }, json_dumps_params={'ensure_ascii': False})
        
    except Exception as e:
        logger.error(f"Error creating time record: {e}", exc_info=True)
        try:
            error_msg = str(e)
            if isinstance(error_msg, bytes):
                error_msg = error_msg.decode('utf-8', errors='replace')
        except Exception:
            error_msg = 'Unknown error occurred'
        
        return JsonResponse({
            'success': False,
            'error': f'Chyba při vytváření záznamu: {error_msg}'
        }, status=500, json_dumps_params={'ensure_ascii': False})
